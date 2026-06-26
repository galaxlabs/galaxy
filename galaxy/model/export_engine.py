import csv
import io
import json
from datetime import UTC, datetime

from galaxy.config import load_site_config
from galaxy.database.connection import get_engine
from galaxy.model.repository import get_doctype
from galaxy.model.document import get_crud_fields
from galaxy.model.document import _quote, _tenant_where
from galaxy.model.permission_engine import apply_field_permissions, apply_data_masks, get_field_restrictions, get_effective_mask_rules
from galaxy.model.repository import get_runtime_meta
from sqlalchemy import text


def _get_engine():
    _, site = load_site_config()
    return get_engine(site)


def _export_filename(doctype: str, fmt: str) -> str:
    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    return f"{doctype}_{ts}.{fmt}"


def _field_names(fields: list[dict]) -> list[str]:
    return ["name"] + [f["fieldname"] for f in fields]


def _query_docs(doctype: str, fields: list[dict], limit: int = 1000, offset: int = 0, search: str = "") -> list[dict]:
    col_names = _field_names(fields)
    quoted_cols = ", ".join(_quote(c) for c in col_names)
    dt = get_doctype(doctype)
    if dt is None:
        return []
    table_name = dt.get("table_name", f"tab{doctype}")
    quoted_table = _quote(table_name)
    tenant_clause, tenant_params = _tenant_where(table_name)
    where_parts = [tenant_clause] if tenant_clause else []
    params = {}
    if search:
        where_parts.append(f"name ILIKE :search")
        params["search"] = f"%{search}%"
    where_sql = " AND ".join(where_parts) if where_parts else "TRUE"
    sql = f"SELECT {quoted_cols} FROM {quoted_table} WHERE {where_sql} ORDER BY name LIMIT :lim OFFSET :off"
    params["lim"] = limit
    params["off"] = offset
    params.update(tenant_params)
    engine = _get_engine()
    with engine.connect() as conn:
        rows = conn.execute(text(sql), params).mappings().all()
    return [dict(r) for r in rows]


def _serialize_value(val, fieldtype: str):
    if val is None:
        return ""
    if fieldtype in ("JSON",):
        if isinstance(val, (dict, list)):
            return json.dumps(val, ensure_ascii=False, default=str)
        return str(val)
    if fieldtype in ("Check",):
        return "Yes" if val else "No"
    return str(val)


def export_csv(doctype: str, fields: list[dict], role: str | None = None, limit: int = 1000, offset: int = 0, search: str = "") -> tuple[str, str, str]:
    rows = _query_docs(doctype, fields, limit, offset, search)
    if role:
        meta = get_runtime_meta(doctype)
        if meta:
            read_fields, _ = get_field_restrictions(meta, role)
            masks = get_effective_mask_rules(meta, role)
            rows = [apply_field_permissions(r, role, read_fields or []) for r in rows]
            rows = [apply_data_masks(r, role, masks) for r in rows]
    col_names = _field_names(fields)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(col_names)
    fieldtype_map = {f["fieldname"]: f["fieldtype"] for f in fields}
    fieldtype_map["name"] = "Data"
    for row in rows:
        writer.writerow([_serialize_value(row.get(c), fieldtype_map.get(c, "Data")) for c in col_names])
    content = buf.getvalue()
    return content, "text/csv", _export_filename(doctype, "csv")


def export_json(doctype: str, fields: list[dict], role: str | None = None, limit: int = 1000, offset: int = 0, search: str = "") -> tuple[str, str, str]:
    rows = _query_docs(doctype, fields, limit, offset, search)
    if role:
        meta = get_runtime_meta(doctype)
        if meta:
            read_fields, _ = get_field_restrictions(meta, role)
            masks = get_effective_mask_rules(meta, role)
            rows = [apply_field_permissions(r, role, read_fields or []) for r in rows]
            rows = [apply_data_masks(r, role, masks) for r in rows]
    content = json.dumps(rows, ensure_ascii=False, default=str, indent=2)
    return content, "application/json", _export_filename(doctype, "json")


def export_excel(doctype: str, fields: list[dict], role: str | None = None, limit: int = 1000, offset: int = 0, search: str = "") -> tuple[bytes, str, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Install with: pip install openpyxl")
    rows = _query_docs(doctype, fields, limit, offset, search)
    if role:
        meta = get_runtime_meta(doctype)
        if meta:
            read_fields, _ = get_field_restrictions(meta, role)
            masks = get_effective_mask_rules(meta, role)
            rows = [apply_field_permissions(r, role, read_fields or []) for r in rows]
            rows = [apply_data_masks(r, role, masks) for r in rows]
    col_names = _field_names(fields)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = doctype[:31]
    header_font = Font(bold=True)
    for ci, name in enumerate(col_names, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = header_font
    fieldtype_map = {f["fieldname"]: f["fieldtype"] for f in fields}
    fieldtype_map["name"] = "Data"
    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(col_names, 1):
            ws.cell(row=ri, column=ci, value=_serialize_value(row.get(col), fieldtype_map.get(col, "Data")))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", _export_filename(doctype, "xlsx")


EXPORT_FORMATS = {
    "csv": export_csv,
    "json": export_json,
    "xlsx": export_excel,
}


def export_docs(doctype: str, fmt: str = "csv", role: str | None = None, limit: int = 1000, offset: int = 0, search: str = ""):
    fmt = fmt.lower()
    if fmt not in EXPORT_FORMATS:
        raise ValueError(f"Unsupported format: {fmt}. Use csv, json, or xlsx.")
    fields = get_crud_fields(doctype)
    if not fields:
        raise ValueError(f"DocType '{doctype}' not found or has no CRUD fields.")
    return EXPORT_FORMATS[fmt](doctype, fields, role=role, limit=limit, offset=offset, search=search)


__all__ = ["export_docs", "export_csv", "export_json", "export_excel"]
